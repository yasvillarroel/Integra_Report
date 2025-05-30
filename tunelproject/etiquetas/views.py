from datetime import datetime # para trabajar con fechas y horas.
from io import BytesIO # para manejar datos binarios en memoria como si fueran un archivo.
from django.utils import timezone # para trabajar con zonas horarias y obtener la hora actual.
import locale
import os # para interactuar con el sistema operativo, como manejar rutas de archivos y directorios.
import tempfile # para trabajar con archivos temporales (creados y eliminados automáticamente).
from django.shortcuts import render, redirect # para renderizar plantillas y 'redirect' para redirigir a otra vista.
from django.contrib.auth import authenticate, login, logout # para autenticar, iniciar sesión y cerrar sesión de usuarios.

from .models import CargarArchivo, Etiqueta # para trabajar con los datos relacionados con este modelo.
from django.db import models # para crear y gestionar modelos de base de datos en Django.
from .models import Etiqueta, Usuario, RolUser # para trabajar con los modelos de la aplicación.

from django.http import HttpResponse, JsonResponse # para crear respuestas HTTP, como devolver archivos (PDF, Excel).
import pandas as pd # para trabajar con estructuras de datos como tablas (especialmente útil para trabajar con Excel y CSV).
from reportlab.lib.styles import getSampleStyleSheet # para obtener un conjunto predefinido de estilos para los elementos de texto en PDF.
from openpyxl import load_workbook # para cargar archivos Excel (.xlsx) y manipularlos en Python.
from openpyxl.styles import PatternFill, Font # para aplicar color a celdas y 'Font' para cambiar la fuente en un archivo Excel.

from reportlab.lib.pagesizes import landscape, legal, inch # como orientación horizontal ('landscape') y tamaño de página legal ('legal').
from reportlab.lib import colors # para trabajar con colores al crear documentos PDF.
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer, Paragraph # para crear documentos PDF, 'Table' para tablas, y 'Paragraph' para párrafos de texto.

from django.contrib.auth.models import User # para trabajar con los usuarios del sistema en Django.
from django.shortcuts import redirect, get_object_or_404 # para redirigir a otras vistas y obtener objetos de la base de datos, respectivamente.
from django.contrib import messages # para mostrar mensajes de éxito, error o advertencia en la interfaz de usuario.

from django.contrib.auth.decorators import login_required # para requerir que el usuario inicie sesión antes de acceder a una vista.

from django.contrib.auth.forms import PasswordChangeForm # para cambiar la contraseña de un usuario.
from django.contrib.auth import update_session_auth_hash # para mantener la sesión activa después de cambiar la contraseña.

from django.contrib.auth.hashers import make_password

from django.core.files.storage import default_storage # para manejar el almacenamiento de archivos, como subir y guardar archivos en el sistema de archivos del servidor.

from unidecode import unidecode


from django.db.models import Sum

from django.core.files.storage import FileSystemStorage
from django.conf import settings


#-------------------------------------------------------------Página de Etiquetas------------------------------------------------------------------
def etiquetas(request):
    label = Etiqueta.objects.all()
    return render(request, "etiquetas.html", {"etiquetas": label})
#----------------------------------------------------------------Página de Login-------------------------------------------------------------------
def login_view(request):
    if request.method == "POST":
        username = request.POST["username"]
        password = request.POST["password"]
        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            return redirect("/index")  # Redirigir al home o la página que prefieras
        else:
            return render(request, "login.html", {"error": "Usuario o contraseña incorrectos"})

    return render(request, "login.html")

def logout_view(request):
    logout(request)
    return redirect("login")
#----------------------------------------------------------------Página de Inicio------------------------------------------------------------------
def index(request):
    
    locale.setlocale(locale.LC_TIME, "es_ES.UTF-8")  # Para sistemas Unix/Linux/Mac
    # Obtener el mes actual
    mes_actual = datetime.now().strftime("%Y-%m")
    
    # Obtener el nombre del mes en texto (Ej: "Febrero")
    nombre_mes = datetime.now().strftime("%B").capitalize()
    
    # Obtener los datos de producción para el mes actual
    datos_produccion_mensual = (
        Etiqueta.objects.filter(DiaProceso2__startswith=mes_actual)
        .values('DestinoProducto')
        .annotate(total_kilos=Sum('TotalKilos')))
    
    # Preparar los datos para el gráfico
    destinos = [dato['DestinoProducto'] for dato in datos_produccion_mensual]
    total_kilos = [float(dato['total_kilos']) for dato in datos_produccion_mensual]  # Asegúrate de que sean números
    
    # Depuración en consola del servidor
    print("Mes actual:", nombre_mes)
    print("Destinos:", destinos)
    print("Total Kilos:", total_kilos)
    
    return render(request, 'index.html', {
        'destinos': destinos,
        'total_kilos': total_kilos,
        'mes': nombre_mes  # Pasamos el nombre del mes a la plantilla
    })
#----------------------------------------------------------------Página de Reportes-----------------------------------------------------------------
def reportes_view(request):
    # Obtener parámetros de filtrado desde la URL
    lote = request.GET.get("lote", "").strip()
    fecha = request.GET.get("fecha", "").strip()
    corte = request.GET.get("corte", "").strip()
    area = request.GET.get("area", "").strip()
    destino_producto = request.GET.get("destino_producto", "").strip()
    calidad =  request.GET.get("calidad", "").strip()
    hora_intervalo = request.GET.get("hora_intervalo", "").strip()
    turno = request.GET.get("turno", "").strip()
    mes_proceso = request.GET.get("mes_proceso", "").strip()

    # Obtener todas las etiquetas inicialmente
    etiquetas = Etiqueta.objects.all()

    # Aplicar filtros si hay valores ingresados
    if lote:
        etiquetas = etiquetas.filter(Lote__icontains=lote)  # Filtrar por lote
    if fecha:
        etiquetas = etiquetas.filter(FechaProduccion=fecha)  # Filtrar por fecha
    if corte:
        # Filtrar "PORCION C/PIEL" o "PORCION S/PIEL" exactamente
        if corte == "PORCION S/PIEL":
            etiquetas = etiquetas.filter(Corte="PORCION S/PIEL")
        elif corte == "PORCION C/PIEL":
            etiquetas = etiquetas.filter(Corte="PORCION C/PIEL")
        else:
            # Si contiene "PORCION C/PIEL" o "PORCION S/PIEL" sin más texto, solo mostrar los cortos
            if corte.startswith("PORCION C/PIEL") or corte.startswith("PORCION S/PIEL"):
                etiquetas = etiquetas.filter(Corte__exact=corte)  # Exact match for these cases
            else:
                # Si no es ninguno de los anteriores, realizar búsqueda por subcadena
                etiquetas = etiquetas.filter(Corte__icontains=corte)
    if area:
        etiquetas = etiquetas.filter(Area=area)
    if destino_producto:
        etiquetas = etiquetas.filter(DestinoProducto=destino_producto)
    if calidad:
        etiquetas = etiquetas.filter(Calidad=calidad)
    if hora_intervalo:
        etiquetas = etiquetas.filter(HoraIntervalo=hora_intervalo)
    if turno:
        etiquetas = etiquetas.filter(Turno=turno)
    if mes_proceso:
        etiquetas = etiquetas.filter(MesProceso=mes_proceso)
        
        
     # Calcular los totales
    total_kilos = round(etiquetas.aggregate(total_kilos=models.Sum('TotalKilos'))['total_kilos'] or 0, 2)
    total_bandejas = etiquetas.aggregate(total_bandejas=models.Sum('TotalBandejas'))['total_bandejas'] or 0

    # Renderizar la vista con los totales
    return render(request, 'reportes.html', {
        'etiquetas': etiquetas,
        'total_kilos': total_kilos,
        'total_bandejas': total_bandejas
    })
#-------------------------------------------------------------Página de Exportar Excel--------------------------------------------------------------
def export_to_excel(request):
    # Obtener parámetros de filtrado desde la URL
    lote = request.GET.get('lote', '').strip()
    fecha = request.GET.get('fecha', '').strip()
    corte = request.GET.get('corte', '').strip()
    area = request.GET.get('area', '').strip()
    destino_producto = request.GET.get('destino_producto', '').strip()
    calidad = request.GET.get('calidad', '').strip()
    hora_intervalo = request.GET.get('hora_intervalo', '').strip()
    turno = request.GET.get('turno', '').strip()
    mes_proceso = request.GET.get('mes_proceso', '').strip()

    # Comenzar con todos los registros
    query = Etiqueta.objects.all()

    # Aplicar filtros solo si los parámetros no están vacíos
    if lote:
        query = query.filter(Lote__icontains=lote)
    if fecha:
        query = query.filter(FechaProduccion=fecha)
    if corte:
        query = query.filter(Corte__icontains=corte)
    if area:
        query = query.filter(Area=area)
    if destino_producto:
        query = query.filter(DestinoProducto=destino_producto)
    if calidad:
        query = query.filter(Calidad=calidad)
    if hora_intervalo:
        query = query.filter(HoraIntervalo=hora_intervalo)
    if turno:
        query = query.filter(Turno=turno)
    if mes_proceso:
        query = query.filter(MesProceso=mes_proceso)

    # Verificar si hay etiquetas después de aplicar los filtros
    if query.count() == 0:
        return HttpResponse("No se encontraron datos que coincidan con los filtros aplicados.", status=400)

    # Convertir los datos filtrados a un DataFrame de pandas
    data = list(query.values())
    df = pd.DataFrame(data)
    
    # Eliminar las columnas no deseadas
    df = df.drop(columns=['FechaProduccion', 'DiaProceso2'])
    
    # Reordenar las columnas
    df = df[['DiaProceso', 'CodigoBarra', 'Lote', 'Area', 'DestinoProducto', 'TotalKilos', 'TotalBandejas', 
             'Corte', 'Calidad', 'Conservacion', 'HoraProceso', 'HoraIntervalo', 'Turno', 'MesProceso']]
    
    # Formatear la columna 'DiaProceso' para mostrar solo la fecha en formato DD-MM-AAAA
    df['DiaProceso'] = df['DiaProceso'].apply(lambda x: x.strftime("%d-%m-%Y") if isinstance(x, datetime) else x)

    # Eliminar la hora de las fechas
    for column in df.columns:
        if pd.api.types.is_datetime64_any_dtype(df[column]):
            df[column] = df[column].dt.date

    # Eliminar la columna 'id' si existe
    if 'id' in df.columns:
        df = df.drop(columns=['id'])

    # Verifica si la columna de fecha tiene zona horaria y la elimina
    for column in df.columns:
        if pd.api.types.is_datetime64_any_dtype(df[column]):
            df[column] = df[column].dt.tz_localize(None)

    # Crear un archivo temporal en disco
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        file_path = tmp.name

    # Convertir las columnas numéricas a tipo float
    df['TotalKilos'] = pd.to_numeric(df['TotalKilos'], errors='coerce')
    df['TotalBandejas'] = pd.to_numeric(df['TotalBandejas'], errors='coerce')

    # Calcular los totales antes de agregar las unidades
    total_kilos = round(df['TotalKilos'].sum(), 2)
    total_bandejas = int(df['TotalBandejas'].sum())  # Cambiado de math.trunc a int()

    # Formatear las columnas con sus unidades
    df['TotalKilos'] = df['TotalKilos'].apply(lambda x: f"{x:.2f} KG" if pd.notna(x) else '')
    df['TotalBandejas'] = df['TotalBandejas'].apply(lambda x: f"{int(x)} UNIDADES" if pd.notna(x) else '')

    # Exportar el DataFrame a Excel
    df.to_excel(file_path, index=False, engine='openpyxl')

    # Cargar el archivo para aplicar formato
    workbook = load_workbook(file_path)
    sheet = workbook.active

    # Aplicar formato a las celdas
    fill = PatternFill(start_color="FAF0E6", end_color="FAF0E6", fill_type="solid")
    fill_header = PatternFill(start_color="FF4500", end_color="FF4500", fill_type="solid")

    # Formato para las celdas de datos
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.fill = fill

    # Formato para los encabezados
    for cell in sheet[1]:
        cell.fill = fill_header
        cell.font = Font(color="FFFFFF", bold=True)

    # Agregar fila de totales
    last_row = sheet.max_row + 1
    sheet[f"A{last_row}"] = "Totales"
    sheet[f"F{last_row}"] = f"{total_kilos:.2f} KG"
    sheet[f"G{last_row}"] = f"{total_bandejas} UNIDADES"
    
    fill_totales = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Color gris claro

    # Formato para la fila de totales
    for cell in sheet[last_row]:
        cell.font = Font(bold=True)
        cell.fill = fill_totales  # Aplicar el color gris

    # Ajustar el ancho de las columnas
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_letter].width = adjusted_width

    # Guardar el archivo
    workbook.save(file_path)

    # Preparar la respuesta
    with open(file_path, 'rb') as f:
        response = HttpResponse(f.read(), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="Reporte_Producción_Tunel.xlsx"'

    # Eliminar archivo temporal
    os.remove(file_path)

    return response
#--------------------------------------------------------------Página de Exportar Pdf---------------------------------------------------------------
def export_to_pdf(request):
    # Obtener parámetros de filtrado desde la URL
    lote = request.GET.get('lote', '').strip()
    fecha = request.GET.get('fecha', '').strip()
    corte = request.GET.get('corte', '').strip()
    area = request.GET.get('area', '').strip()
    destino_producto = request.GET.get('destino_producto', '').strip()
    calidad = request.GET.get('calidad', '').strip()
    hora_intervalo = request.GET.get('hora_intervalo', '').strip()
    turno = request.GET.get('turno', '').strip()
    mes_proceso = request.GET.get('mes_proceso', '').strip()

    # Comenzar con todos los registros
    query = Etiqueta.objects.all()

    # Aplicar filtros solo si los parámetros no están vacíos
    if lote:
        query = query.filter(Lote__icontains=lote)
    if fecha:
        query = query.filter(FechaProduccion=fecha)
    if corte:
        query = query.filter(Corte__icontains=corte)
    if area:
        query = query.filter(Area=area)
    if destino_producto:
        query = query.filter(DestinoProducto=destino_producto)
    if calidad:
        query = query.filter(Calidad=calidad)
    if hora_intervalo:
        query = query.filter(HoraIntervalo=hora_intervalo)
    if turno:
        query = query.filter(Turno=turno)
    if mes_proceso:
        query = query.filter(MesProceso=mes_proceso)

    # Verificar los filtros aplicados
    print("Consulta generada:", query.query)  # Imprime la consulta SQL generada

    # Verificar si hay etiquetas después de aplicar los filtros
    if query.count() == 0:
        return HttpResponse("No se encontraron datos que coincidan con los filtros aplicados.", status=400)

    # Crear la respuesta para el PDF
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="pdf_exportado.pdf"'

    # Crear el objeto PDF
    buffer = BytesIO()

    # Crear el documento con tamaño Legal (8.5 x 14 pulgadas) y orientación horizontal
    doc = SimpleDocTemplate(buffer, pagesize=landscape(legal))  # Uso de landscape para orientación horizontal
    
    # Obtener los datos del modelo
    data = query.values()
    
    # Encabezados de la tabla con el orden solicitado
    headers = ["Día Proceso", "Código de Barra", "Lote", "Área", "Destino Producto", "Total Kilos", "Total Bandejas", "Corte", "Calidad", "Conservación", "Hora Proceso", "Hora Intervalo", "Turno", "Mes Proceso"]

    # Formatear la columna 'DiaProceso' para mostrar solo la fecha en formato DD-MM-AAAA
    table_data = [headers]
    for item in data:
        dia_proceso = item["DiaProceso"]
        if isinstance(dia_proceso, datetime):
            dia_proceso = dia_proceso.strftime("%d-%m-%Y")
        
        total_kilos = f"{item['TotalKilos']} KG"  # Agregar la unidad de medida KG
        total_bandejas = f"{item['TotalBandejas']} UNIDADES"  # Agregar la unidad de medida UNIDADES
    
        table_data.append([
            dia_proceso,
            item["CodigoBarra"],
            item["Lote"],
            item["Area"],
            item["DestinoProducto"],
            total_kilos,  # Incluir total_kilos con la unidad
            total_bandejas,  # Incluir total_bandejas con la unidad
            item["Corte"],
            item["Calidad"],
            item["Conservacion"],
            item["HoraProceso"],
            item["HoraIntervalo"],
            item["Turno"],
            item["MesProceso"]
        ])
        
    # Calcular los totales solo una vez, después de procesar todas las filas
    total_kilos = round(query.aggregate(total_kilos=models.Sum('TotalKilos'))['total_kilos'] or 0, 2)
    total_bandejas = query.aggregate(total_bandejas=models.Sum('TotalBandejas'))['total_bandejas'] or 0

    # Agregar las unidades de medida al final
    total_kilos_str = f"{total_kilos} KG"
    total_bandejas_str = f"{total_bandejas} UNIDADES"

    # Agregar los totales al final de la tabla, una vez fuera del bucle
    table_data.append(["TOTALES", "", "", "", "", total_kilos_str, total_bandejas_str, "", "", "", "", "", ""])

    # Crear un objeto de estilo para el título
    styles = getSampleStyleSheet()
    title_style = styles['Title']
    title_style.fontName = 'Helvetica-Bold'
    title_style.fontSize = 18

    # Crear el título
    title = Paragraph("Reporte de Producción Tunel Continuo", title_style)

        
    custom_size = (16 * inch, 13 * inch)  # Multiplicamos por inch para obtener el tamaño en pulgadas
        
    doc = SimpleDocTemplate(buffer, pagesize=custom_size)  # Usa el tamaño personalizado

    # Crear la tabla
    table = Table(table_data)
         
    # Estilos de la tabla
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.orangered),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.linen),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),  # Fondo gris para la fila de totales
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),  # Negrita para los totales
    ]))

    # Crear un espacio antes de la tabla
    spacer = Spacer(1, 12)

    # Usamos el método `build` de SimpleDocTemplate para crear el PDF con el título y la tabla
    doc.build([title, spacer, table])

    # Regresar el archivo como una respuesta HTTP
    buffer.seek(0)
    response = HttpResponse(buffer, content_type="application/pdf")
    response['Content-Disposition'] = 'attachment; filename="Reporte_Producción_Tunel.pdf"'
    
    return response
#---------------------------------------------------------------Página de Perfil------------------------------------------------------------------
@login_required
def perfil_view(request):
    # Obtener el perfil personalizado del usuario autenticado
    try:
        perfil = Usuario.objects.get(email_user=request.user.email)
    except Usuario.DoesNotExist:
        messages.error(request, "No se encontró el perfil asociado.")
        return render(request, 'perfil.html')

    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        apellido = request.POST.get('apellido')
        email = request.POST.get('email')

        if nombre and apellido and email:
            request.user.email = email
            request.user.save()

            perfil.nombre_user = nombre
            perfil.apellido_user = apellido
            perfil.email_user = email
            perfil.save()

            messages.success(request, "Perfil actualizado correctamente.")
            return redirect('perfil')
        else:
            messages.error(request, "Todos los campos son obligatorios.")

    return render(request, 'perfil.html', {'perfil': perfil})
#------------------------------------------------------------Página de Cambiar Contraseña------------------------------------------------------------
@login_required
def contraseña_view(request):
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            form.save()
            update_session_auth_hash(request, form.user)  # Mantener la sesión activa
            messages.success(request, 'Contraseña modificada correctamente')
            return redirect('cambiar-contraseña')  # Redirige a la página de perfil después de cambiar la contraseña
        else:
            messages.error(request, 'La contraseña no es válida. Inténtalo de nuevo.')
            print(form.errors)  # Mostrar errores en consola

    else:
        form = PasswordChangeForm(request.user)

    return render(request, 'cambiar-contraseña.html')
#------------------------------------------------------------Página de Usuarios--------------------------------------------------------------------

from .models import Usuario, RolUser, CargarArchivo, Etiqueta

def usuarios_view(request):
    usuarios_django = User.objects.all()
    usuarios_custom = Usuario.objects.all()
    roles = RolUser.objects.all()

    lista_usuarios = []

    for user in usuarios_django:
        perfil = usuarios_custom.filter(email_user=user.email).first()  # ✅ relación confiable por email
        lista_usuarios.append({
            'django': user,
            'perfil': perfil
        })

    return render(request, 'usuarios.html', {
        'usuarios': lista_usuarios,
        'roles': roles
    })



#------------------------------------------------------------ Agregar Usuario --------------------------------------------------------------------
def agregar_usuario(request):
    if request.method == "POST":
        username = request.POST.get("username")
        apellido = request.POST.get("apellido")
        email = request.POST.get("email")
        password = request.POST.get("password")
        id_rol = request.POST.get("user_type")

        if User.objects.filter(username=username).exists():
            messages.error(request, "El usuario ya existe.")
            return redirect("usuarios")

        if Usuario.objects.filter(email_user=email).exists():
            messages.error(request, "Ese correo ya está registrado.")
            return redirect("usuarios")

        is_superuser = int(id_rol) == 21

        # Crear en auth_user
        django_user = User.objects.create(
            username=username,
            email=email,
            password=make_password(password),
            is_superuser=is_superuser,
            is_staff=is_superuser
        )

        # Crear en la tabla personalizada
        Usuario.objects.create(
            nombre_user=username,
            apellido_user=apellido,
            email_user=email,
            pass_user=make_password(password),
            id_rol=RolUser.objects.get(id_rol=id_rol)
        )

        messages.success(request, "Usuario agregado con éxito.")
        return redirect("usuarios")


    #---------------------------------------------------------Eliminar Usuario---------------------------------------------------------------------
def eliminar_usuario(request, usuario_id):
    usuario = get_object_or_404(User, id=usuario_id)

    # Eliminar también en la tabla personalizada si existe
    try:
        usuario_personal = Usuario.objects.get(nombre_user=usuario.username)
        usuario_personal.delete()
    except Usuario.DoesNotExist:
        pass  # Si no existe, no pasa nada

    usuario.delete()
    messages.success(request, "Usuario eliminado correctamente.")
    return redirect('usuarios')
    #---------------------------------------------------------Editar Usuario-----------------------------------------------------------------------
def editar_usuario(request, usuario_id):
    usuario = get_object_or_404(User, id=usuario_id)

    if request.method == 'POST':
        nuevo_nombre = request.POST.get('username')
        nuevo_apellido = request.POST.get('apellido')
        nuevo_email = request.POST.get('email')
        nuevo_rol_id = request.POST.get('user_type')

        # 🔍 Buscar perfil primero antes de cambiar username
        try:
            perfil = Usuario.objects.get(nombre_user=usuario.username)
        except Usuario.DoesNotExist:
            perfil = None

        # Verificar si el correo ya existe en otro usuario
        if Usuario.objects.exclude(nombre_user=usuario.username).filter(email_user=nuevo_email).exists():
            messages.error(request, "El correo ya está en uso por otro usuario.")
            return redirect('usuarios')

        # Actualizar auth_user
        usuario.username = nuevo_nombre
        usuario.email = nuevo_email
        usuario.save()

        # Actualizar tabla personalizada
        if perfil:
            perfil.nombre_user = nuevo_nombre
            perfil.apellido_user = nuevo_apellido
            perfil.email_user = nuevo_email
            perfil.id_rol = RolUser.objects.get(id_rol=nuevo_rol_id)
            perfil.save()

        messages.success(request, "Usuario actualizado correctamente.")
        return redirect('usuarios')

    return redirect('usuarios')


#-----------------------------------------------------------Página de Estadísticas-------------------------------------------------------------------
def estadisticas_view(request):
    datos_produccion = (
        Etiqueta.objects.values('DiaProceso2', 'DestinoProducto')
        .annotate(total_kilos=Sum('TotalKilos'))
        .order_by('DiaProceso2', 'DestinoProducto')
    )

    fechas = sorted(set(dato["DiaProceso2"].strftime("%Y-%m-%d") for dato in datos_produccion))
    destinos = sorted(set(dato["DestinoProducto"] for dato in datos_produccion))

    datos_destinos = {destino: [0] * len(fechas) for destino in destinos}

    for dato in datos_produccion:
        fecha = dato["DiaProceso2"].strftime("%Y-%m-%d")
        destino = dato["DestinoProducto"]
        index = fechas.index(fecha)
        datos_destinos[destino][index] = float(dato["total_kilos"])

    return render(request, "estadisticas.html", {
        "fechas": fechas,
        "destinos": destinos,
        "datos_destinos": datos_destinos
    })


#---------------------------------------------------------------Pruebas gráficos-----------------------------------------------------------------
def graficos_view(request):
    datos_produccion = (
        Etiqueta.objects.values('DiaProceso2', 'DestinoProducto')
        .annotate(total_kilos=Sum('TotalKilos'))
        .order_by('DiaProceso2', 'DestinoProducto')
    )

    fechas = sorted(set(dato["DiaProceso2"].strftime("%Y-%m-%d") for dato in datos_produccion))
    destinos = sorted(set(dato["DestinoProducto"] for dato in datos_produccion))

    datos_destinos = {destino: [0] * len(fechas) for destino in destinos}

    for dato in datos_produccion:
        fecha = dato["DiaProceso2"].strftime("%Y-%m-%d")
        destino = dato["DestinoProducto"]
        index = fechas.index(fecha)
        datos_destinos[destino][index] = float(dato["total_kilos"])

    return render(request, "graficos.html", {
        "fechas": fechas,
        "destinos": destinos,
        "datos_destinos": datos_destinos
    })
    
#---------------------------------------------------------------Importar excel actualizado-----------------------------------------------------------------
def importar_base_de_datos(request):
    if request.method == "POST" and request.FILES['archivo_excel']:
        archivo_excel = request.FILES['archivo_excel']  # Obtener el archivo subido

        try:
            # Leer el archivo Excel
            df = pd.read_excel(archivo_excel)

            # Convertir las columnas de fecha
            df['DiaProceso'] = pd.to_datetime(df['DiaProceso'], errors='coerce', format='%Y-%m-%d')
            df['DiaProceso2'] = pd.to_datetime(df['DiaProceso2'], errors='coerce', format='%Y-%m-%d')
            df['FechaProduccion'] = pd.to_datetime(df['FechaProduccion'], errors='coerce', format='%Y-%m-%d')

            # Verificar que las fechas se hayan convertido correctamente
            if df['DiaProceso2'].isna().any():
                messages.error(request, "Algunas fechas en 'DiaProceso2' no pudieron ser convertidas. Verifica el formato en el archivo Excel.")
                return redirect('importar_base_de_datos')

            # Iterar sobre las filas del DataFrame
            for index, row in df.iterrows():
                if not Etiqueta.objects.filter(CodigoBarra=row['CodigoBarra'], FechaProduccion=row['FechaProduccion']).exists():
                    Etiqueta.objects.create(
                        CodigoBarra=row['CodigoBarra'],
                        FechaProduccion=row['FechaProduccion'],
                        DiaProceso2=row['DiaProceso2'].date(),  # Convertir a date si es necesario
                        Lote=row['Lote'],
                        Area=row['Area'],
                        DestinoProducto=row['DestinoProducto'],
                        TotalKilos=row['TotalKilos'],
                        Corte=row['Corte'],
                        Calidad=row['Calidad'],
                        Conservacion=row['Conservacion'],
                        HoraProceso=row['HoraProceso'],
                        HoraIntervalo=row['HoraIntervalo'],
                        Turno=row['Turno'],
                        MesProceso=row['MesProceso'],
                    )

            # Mensaje de éxito
            messages.success(request, "Datos importados correctamente.")
            return redirect('reportes')

        except Exception as e:
            messages.error(request, f"Hubo un error al procesar el archivo: {str(e)}")
            return redirect('importar_base_de_datos')

    else:
        return render(request, 'reportes.html')
#---------------------------------------------------------------Importar excel actualizado-----------------------------------------------------------------

def cargar_archivo_view(request):
    if request.method == 'POST':
        archivo = request.FILES.get('archivo_excel')

        # Buscar el usuario personalizado
        try:
            usuario_personalizado = Usuario.objects.get(email_user=request.user.email)
        except Usuario.DoesNotExist:
            messages.error(request, "Tu perfil no está registrado en la base de datos.")
            return redirect('cargar_archivo')

        if archivo and archivo.name.endswith('.xlsx'):
            archivo_registrado = CargarArchivo.objects.create(
                nombre_archivo=archivo.name,
                id_user=usuario_personalizado
            )

            fs = FileSystemStorage(location=os.path.join(settings.MEDIA_ROOT, 'archivos_excel'))
            filename = fs.save(archivo.name, archivo)
            archivo_path = fs.path(filename)

            try:
                df = pd.read_excel(archivo_path)

                # ✅ Limpieza definitiva de nombres de columnas
                df.columns = [
                    unidecode(col).strip().lower().replace(" ", "").replace("-", "")
                    for col in df.columns
                ]

                print("✅ Columnas limpias y normalizadas:", df.columns.tolist())

                columnas_obligatorias = [
                    'diaproceso', 'codigobarra', 'lote', 'area', 'destinoproducto',
                    'totalkilos', 'totalbandejas', 'horaproceso', 'horaintervalo',
                    'turno', 'mesproceso'
                ]

                faltantes = [col for col in columnas_obligatorias if col not in df.columns]
                if faltantes:
                    messages.error(request, f"❌ Columnas faltantes en el Excel: {', '.join(faltantes)}")
                    print("❌ Columnas faltantes:", faltantes)
                    return redirect('cargar_archivo')

                # ✅ Inserción segura
                for _, row in df.iterrows():
                    try:
                        
                        # Verifica si ya existe ese código de barra
                        if Etiqueta.objects.filter(CodigoBarra=str(row['codigobarra'])).exists():
                            print(f"⚠️ Código ya existente, se omite: {row['codigobarra']}")
                            continue
                        
                        Etiqueta.objects.create(
                            DiaProceso=pd.to_datetime(row['diaproceso']).date(),
                            CodigoBarra=str(row['codigobarra']),
                            Lote=str(row['lote']),
                            Area=str(row['area']),
                            DestinoProducto=str(row['destinoproducto']),
                            TotalKilos=float(row['totalkilos']),
                            TotalBandejas=int(row['totalbandejas']),
                            Corte=str(row.get('corte', '')),
                            Calidad=str(row.get('calidad', '')),
                            Conservacion=str(row.get('conservacion', '')),
                            HoraProceso=pd.to_datetime(row['horaproceso']).time(),
                            HoraIntervalo=str(row['horaintervalo']),
                            Turno=str(row['turno']),
                            MesProceso=str(row['mesproceso']),
                            FechaProduccion=pd.to_datetime(row.get('fechaproduccion')).date()
                                if pd.notnull(row.get('fechaproduccion')) else None,
                            DiaProceso2=pd.to_datetime(row.get('diaproceso2')).date()
                                if pd.notnull(row.get('diaproceso2')) else None,
                            archivo_origen=archivo_registrado
                        )
                    except Exception as fila_error:
                        print(f"❌ Error al insertar fila: {fila_error}")
                        continue

                messages.success(request, "Archivo cargado y datos procesados exitosamente.")
            except Exception as e:
                print("❌ ERROR AL PROCESAR:", e)
                messages.error(request, f"No se pudo procesar el archivo: {str(e)}")
                return redirect('cargar_archivo')
        else:
            messages.error(request, "Formato no válido. Solo se permiten archivos .xlsx.")
        return redirect('cargar_archivo')

    archivos = CargarArchivo.objects.order_by('-fecha_subida')[:10]
    return render(request, 'cargar_archivo.html', {'archivos': archivos})